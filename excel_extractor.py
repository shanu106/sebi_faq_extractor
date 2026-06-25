"""
Excel URL Extractor and PDF FAQ Scraper Utilities
"""

import re
import csv
import logging
from io import BytesIO, StringIO
from typing import List, Dict, Any, Set, Optional
import requests
import openpyxl
import PyPDF2

logger = logging.getLogger(__name__)

def extract_pdf_links_from_excel(file_contents: bytes) -> List[str]:
    """
    Reads an Excel file and extracts all unique PDF URLs.
    Looks at cell text content and cell hyperlinks.
    """
    pdf_links: List[str] = []
    seen_links: Set[str] = set()

    # Regular expressions for finding HTTP/HTTPS URLs
    pdf_pattern = re.compile(r'https?://[^\s"\'<>]+?\.pdf(?:[?#][^\s"\'<>]*)?', re.IGNORECASE)
    general_url_pattern = re.compile(r'https?://[^\s"\'<>]+', re.IGNORECASE)

    try:
        wb = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows():
                for cell in row:
                    # 1. Check cell hyperlinks
                    if cell.hyperlink and cell.hyperlink.target:
                        url = cell.hyperlink.target
                        if '.pdf' in url.lower() and url not in seen_links:
                            seen_links.add(url)
                            pdf_links.append(url)
                            continue

                    # 2. Check cell text value
                    val = cell.value
                    if val and isinstance(val, str):
                        # Extract any matching PDF URL
                        matches = pdf_pattern.findall(val)
                        for url in matches:
                            if url not in seen_links:
                                seen_links.add(url)
                                pdf_links.append(url)
                        
                        # Fallback: scan any URLs and check for '.pdf'
                        if not matches:
                            urls = general_url_pattern.findall(val)
                            for url in urls:
                                if '.pdf' in url.lower() and url not in seen_links:
                                    seen_links.add(url)
                                    pdf_links.append(url)

    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise ValueError(f"Failed to parse Excel workbook: {str(e)}")

    logger.info(f"Extracted {len(pdf_links)} unique PDF links from Excel")
    return pdf_links


def extract_pdf_links_from_csv(file_contents: bytes) -> List[str]:
    """
    Reads a CSV file and extracts all unique PDF URLs.
    """
    pdf_links: List[str] = []
    seen_links: Set[str] = set()

    # Regular expressions for finding HTTP/HTTPS URLs
    pdf_pattern = re.compile(r'https?://[^\s"\'<>]+?\.pdf(?:[?#][^\s"\'<>]*)?', re.IGNORECASE)
    general_url_pattern = re.compile(r'https?://[^\s"\'<>]+', re.IGNORECASE)

    try:
        # Decode contents
        try:
            text = file_contents.decode('utf-8')
        except UnicodeDecodeError:
            text = file_contents.decode('latin-1')

        reader = csv.reader(StringIO(text))
        for row in reader:
            for val in row:
                if val:
                    # Extract any matching PDF URL
                    matches = pdf_pattern.findall(val)
                    for url in matches:
                        if url not in seen_links:
                            seen_links.add(url)
                            pdf_links.append(url)
                    
                    # Fallback: scan any URLs and check for '.pdf'
                    if not matches:
                        urls = general_url_pattern.findall(val)
                        for url in urls:
                            if '.pdf' in url.lower() and url not in seen_links:
                                seen_links.add(url)
                                pdf_links.append(url)

    except Exception as e:
        logger.error(f"Error parsing CSV file: {e}")
        raise ValueError(f"Failed to parse CSV file: {str(e)}")

    logger.info(f"Extracted {len(pdf_links)} unique PDF links from CSV")
    return pdf_links


def scrape_faqs_from_pdf_bytes(pdf_bytes: bytes, source_url: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Multi-strategy FAQ extraction from PDF bytes.
    
    Strategies (applied in order, results merged):
      1. Numbered questions: Lines starting with N. / QN. with gap-tolerant sequencing
      2. Bold text detection: Uses pdfplumber font metadata to find bold question lines
      3. Question-mark detection: Lines ending with '?' even without serial numbers
      4. Regex fallback: Pattern-matching Q/A markers for unstructured PDFs
    """

    # =====================================================================
    # STEP 1: TEXT EXTRACTION (pdfplumber primary, PyPDF2 fallback)
    # =====================================================================
    bold_line_texts: Set[str] = set()  # Set of line-text strings detected as bold
    text = ""

    # --- Try pdfplumber first (better text extraction + bold detection) ---
    try:
        import pdfplumber
        from collections import defaultdict

        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"

                # Detect bold lines from character-level font info
                try:
                    chars = page.chars
                    if chars:
                        y_groups: Dict[int, list] = defaultdict(list)
                        for c in chars:
                            y_key = round(c['top'])
                            y_groups[y_key].append(c)

                        for y_key in sorted(y_groups.keys()):
                            group_chars = y_groups[y_key]
                            line_text = ''.join(c['text'] for c in group_chars).strip()
                            if not line_text or len(line_text) < 5:
                                continue
                            non_space = [c for c in group_chars if c['text'].strip()]
                            if non_space:
                                bold_count = sum(
                                    1 for c in non_space
                                    if any(w in c.get('fontname', '').lower() for w in ['bold', 'black', 'heavy', 'semibold', 'demibold'])
                                )
                                if bold_count > len(non_space) * 0.5:
                                    bold_line_texts.add(line_text)
                except Exception:
                    pass  # Bold detection is best-effort
    except ImportError:
        logger.info("pdfplumber not available, falling back to PyPDF2 (no bold detection)")
    except Exception as e:
        logger.warning(f"pdfplumber extraction failed ({e}), falling back to PyPDF2")

    # --- Fallback to PyPDF2 if pdfplumber yielded nothing ---
    if not text.strip():
        try:
            pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_bytes))
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except Exception as e:
            logger.error(f"Failed to parse PDF pages: {e}")
            raise RuntimeError(f"PDF parsing error: {str(e)}")

    if not text.strip():
        logger.warning("No text could be extracted from PDF")
        return []

    # =====================================================================
    # STEP 2: HELPER FUNCTIONS
    # =====================================================================
    faqs: List[Dict[str, Any]] = []
    seen_questions: Set[str] = set()

    def clean_text(t: str) -> str:
        t = ' '.join(t.split())
        t = t.replace('&nbsp;', ' ').replace('&amp;', '&')
        return t.strip()

    def clean_question_text(q: str) -> str:
        q_clean = q.strip()
        # Remove footnote junk at start if present
        match = re.search(r'\b(\d+)\.\s+([A-Z])', q_clean)
        if match:
            start_idx = match.start()
            if start_idx > 0:
                q_clean = q_clean[start_idx:]
        # Strip leading Q/number markers (e.g. "Q1.", "1.", "1.1", "Question 1.")
        q_clean = re.sub(r'^\s*(?:Q\d+\.?|Q\.?\s*\d+\.?|\d+(?:\.\d+)?\.?|[Qq]uestion\s*\d+\.?)\s*', '', q_clean)
        return q_clean.strip()

    def clean_answer_text(a: str) -> str:
        a_clean = a.strip()
        # Strip leading footnote numbers
        a_clean = re.sub(r'^\s*\d+\s+([A-Z])', r'\1', a_clean)
        return a_clean.strip()

    def extract_question_number(line_str: str) -> Optional[int]:
        """Extract a leading question number from a line (e.g. '1.', 'Q1.', '1.1', '(1)', '1)')."""
        match = re.match(r'^\s*(?:[Qq](?:uestion)?\.?\s*)?\(?(\d+)\)?(?:[\.:\s\)-]+|$)', line_str)
        if match:
            num = int(match.group(1))
            if num < 500:  # Restrict to avoid matching large numbers or years
                return num
        return None

    def has_question_word(text_str: str) -> bool:
        """Check if text contains question indicator words (using word boundaries)."""
        lower = re.sub(r'^\s*\d+[\.\)\s]+', '', text_str).strip().lower()
        pattern = r'\b(what|how|why|when|where|who|which|can|does|do|is|are|shall|should|will|whether|has|have|had|did|could|would|may|if|whoever|whom)\b'
        return bool(re.search(pattern, lower))

    def is_bold_line(line_text: str) -> bool:
        """Check if a cleaned line matches any detected bold text from pdfplumber."""
        if not bold_line_texts:
            return False
        clean = line_text.strip()
        if len(clean) < 5:
            return False
        for bt in bold_line_texts:
            # Fuzzy match: the cleaned line may have slightly different whitespace
            # than the bold text from pdfplumber, so check containment both ways
            if clean == bt:
                return True
            # Normalize whitespace for comparison
            clean_norm = ' '.join(clean.split())
            bt_norm = ' '.join(bt.split())
            if clean_norm == bt_norm:
                return True
            if len(clean_norm) > 15 and (clean_norm in bt_norm or bt_norm in clean_norm):
                return True
        return False

    def is_likely_question_start(line: str) -> bool:
        """
        Determine if a line is likely the start of a FAQ question.
        Uses multiple signals: numbering, question mark, question words, bold text.
        """
        has_qmark = '?' in line
        has_num = extract_question_number(line) is not None
        has_qword = has_question_word(line)
        bold = is_bold_line(line)

        # Strong: numbered line + question mark
        if has_num and has_qmark:
            return True
        # Strong: bold text + (question mark OR question word)
        if bold and (has_qmark or has_qword):
            return True
        # Moderate: question mark + question word (for unnumbered questions)
        if has_qmark and has_qword and len(line.strip()) > 20:
            return True
        return False

    # =====================================================================
    # STEP 3: LINE CLEANING (remove headers/footers, handle page breaks)
    # =====================================================================
    raw_lines = [line.strip() for line in text.split('\n')]
    lines: List[str] = []

    header_footer_patterns = [
        r'SEBI\s+FAQs?\s+on\s+.*?\s+Page\s+\d+\s+of\s+\d+',
        r'^Page\s+\d+\s+of\s+\d+\s*$',
        r'^\d+\s+of\s+\d+$',
        r'^Frequently\s+Asked\s+Questions?\s+on\s+.*',
        r'^Frequently\s+Asked\s+Questions?\s*$',
        r'^FAQs?\s*$',
    ]

    for line in raw_lines:
        # Skip pure header/footer lines
        is_hf = False
        for pat in header_footer_patterns:
            if re.search(pat, line, re.IGNORECASE):
                is_hf = True
                break

        # Handle lines where a page header is prepended to content
        # e.g. "Page 3 of 10 September 24, 2024 12. What is..."
        header_match = re.search(
            r'^(.*?Page\s+\d+\s+of\s+\d+)\s*(?:[A-Za-z]+\s+\d+\s*,\s*\d+)?\s*',
            line, re.IGNORECASE
        )
        if header_match:
            cleaned = line[header_match.end():].strip()
            if cleaned:
                lines.append(cleaned)
            continue

        if not is_hf:
            lines.append(line)

    if bold_line_texts:
        logger.info(f"Bold detection: found {len(bold_line_texts)} bold text segments in PDF")

    # =====================================================================
    # STEP 4: MULTI-STRATEGY QUESTION BOUNDARY DETECTION
    # =====================================================================

    # --- Strategy 1: Numbered questions with gap-tolerant sequencing ---
    numbered_candidates: List[tuple] = []  # (line_index, question_number)
    for idx, line in enumerate(lines):
        num = extract_question_number(line)
        if num is not None and num >= 1:
            # Check for question mark on this line or within next few lines
            # (stop scanning if we hit another numbered line first)
            has_qmark_nearby = False
            for look_idx in range(idx, min(len(lines), idx + 6)):
                if '?' in lines[look_idx]:
                    has_qmark_nearby = True
                    break
                if look_idx > idx and extract_question_number(lines[look_idx]) is not None:
                    break  # Hit another numbered line — stop

            if has_qmark_nearby:
                numbered_candidates.append((idx, num))

    # Build monotonically increasing sequence with gap tolerance (MAX_GAP)
    question_start_indices: List[int] = []
    if numbered_candidates:
        MAX_GAP = 5  # Allow skipping up to 5 missing question numbers
        last_num = 0
        for (idx, num) in numbered_candidates:
            if num > last_num and num <= last_num + MAX_GAP:
                question_start_indices.append(idx)
                last_num = num
            # If num resets to a small value while we're deep in the sequence,
            # it's answer sub-numbering (e.g. "1. Brokerage charges..." inside Q20)
            # — just skip it silently

    # --- Strategy 2: Bold text questions not already in numbered set ---
    bold_boundaries: List[int] = []
    numbered_set = set(question_start_indices)
    if bold_line_texts:
        for idx, line in enumerate(lines):
            if idx in numbered_set:
                continue
            is_bold = is_bold_line(line)
            if is_bold:
                is_candidate = False
                stripped = line.strip()
                if has_question_word(stripped) or '?' in stripped:
                    is_candidate = True
                elif len(stripped) >= 10 and len(stripped) <= 150:
                    # Allow standalone bold lines as topic headings/questions
                    is_candidate = True

                if is_candidate:
                    # Guard: Check if the line is part of a multi-line numbered question
                    inside_question_part = False
                    for q_start in question_start_indices:
                        if q_start <= idx:
                            q_end = next((s for s in question_start_indices if s > q_start), len(lines))
                            if idx < q_end:
                                block_lines = lines[q_start:q_end]
                                qmark_line_idx = -1
                                for k in range(min(len(block_lines), 5)):
                                    if '?' in block_lines[k]:
                                        qmark_line_idx = k
                                        break
                                if qmark_line_idx != -1 and idx <= q_start + qmark_line_idx:
                                    inside_question_part = True
                                break
                    if inside_question_part:
                        continue

                    # Make sure this bold line isn't deeply inside an existing Q&A block
                    # (i.e. it's not just a bold sub-heading within an answer)
                    inside_existing = False
                    for qi, q_start in enumerate(question_start_indices):
                        q_end = question_start_indices[qi + 1] if qi + 1 < len(question_start_indices) else len(lines)
                        if q_start < idx < q_end:
                            # It's inside an existing block — only add if it has strong signals
                            if '?' in line and has_question_word(line):
                                # Strong enough to be a standalone question within a block
                                pass
                            else:
                                inside_existing = True
                            break
                    if not inside_existing:
                        bold_boundaries.append(idx)

    # --- Strategy 3: Standalone question-mark lines not already detected ---
    qmark_boundaries: List[int] = []
    all_so_far = set(question_start_indices) | set(bold_boundaries)
    for idx, line in enumerate(lines):
        if idx in all_so_far:
            continue
        stripped = line.strip()
        # Relax to match any standalone line ending with a question mark
        if stripped.endswith('?') and len(stripped) >= 10:
            # Guard: Check if the line is part of a multi-line numbered question
            inside_question_part = False
            for q_start in question_start_indices:
                if q_start <= idx:
                    q_end = next((s for s in question_start_indices if s > q_start), len(lines))
                    if idx < q_end:
                        block_lines = lines[q_start:q_end]
                        qmark_line_idx = -1
                        for k in range(min(len(block_lines), 5)):
                            if '?' in block_lines[k]:
                                qmark_line_idx = k
                                break
                        if qmark_line_idx != -1 and idx <= q_start + qmark_line_idx:
                            inside_question_part = True
                        break
            if inside_question_part:
                continue

            # Verify it's not a sub-question inside an existing answer
            # by checking if the NEXT significant content looks like an answer (no '?')
            is_answer_subq = False
            for qi, q_start in enumerate(question_start_indices):
                q_end = question_start_indices[qi + 1] if qi + 1 < len(question_start_indices) else len(lines)
                if q_start < idx < q_end:
                    # Check if this looks like a distinct Q&A (has answer content after it)
                    remaining_lines = q_end - idx - 1
                    if remaining_lines < 2:
                        is_answer_subq = True  # Too close to next boundary
                    break
            if not is_answer_subq:
                qmark_boundaries.append(idx)

    # --- Merge all boundaries ---
    all_boundaries = sorted(set(question_start_indices + bold_boundaries + qmark_boundaries))

    logger.info(
        f"PDF extraction: {len(all_boundaries)} question boundaries found "
        f"({len(question_start_indices)} numbered, {len(bold_boundaries)} bold, "
        f"{len(qmark_boundaries)} question-mark) in {len(lines)} lines"
    )

    # =====================================================================
    # STEP 5: EXTRACT Q&A PAIRS FROM BOUNDARIES
    # =====================================================================
    if all_boundaries:
        for i in range(len(all_boundaries)):
            start_idx = all_boundaries[i]
            end_idx = all_boundaries[i + 1] if i + 1 < len(all_boundaries) else len(lines)

            block_lines = lines[start_idx:end_idx]
            if not block_lines:
                continue

            # ----------------------------------------------------------
            # Find where the question ends: use the FIRST '?' in the
            # first few lines (not the last). This prevents sub-questions
            # inside the answer from being absorbed into the question text.
            # ----------------------------------------------------------
            qmark_line_idx = -1
            for k in range(min(len(block_lines), 5)):
                if '?' in block_lines[k]:
                    qmark_line_idx = k
                    break  # FIRST '?' = end of question

            if qmark_line_idx != -1:
                target_line = block_lines[qmark_line_idx]
                # Find the FIRST '?' in the target line that completes the question
                split_idx = target_line.index('?')
                q_line_end = target_line[:split_idx + 1]
                a_line_start = target_line[split_idx + 1:].strip()

                # Assemble question text (lines before qmark + qmark line up to '?')
                q_text = " ".join(block_lines[:qmark_line_idx]) + " " + q_line_end

                # Assemble answer text (rest of qmark line + remaining lines)
                a_text_parts = []
                if a_line_start:
                    a_text_parts.append(a_line_start)
                a_text_parts.extend(block_lines[qmark_line_idx + 1:])
            else:
                # No question mark found — treat first line as question
                q_text = block_lines[0]
                a_text_parts = block_lines[1:]

            q_cleaned = clean_text(clean_question_text(q_text))

            # Filter residual page numbers / headers from answer text parts BEFORE joining
            filtered_parts = []
            for part in a_text_parts:
                part_str = part.strip()
                if re.search(r'Page \d+ of \d+', part_str, re.I):
                    continue
                if re.search(r'^\d+ of \d+$', part_str):
                    continue
                if re.match(r'^\s*SEBI\s+FAQs?\s*$', part_str, re.I) or re.match(r'^\s*FAQs?\s+on\s+.*$', part_str, re.I):
                    continue
                filtered_parts.append(part_str)
            a_cleaned = clean_text(clean_answer_text(" ".join(filtered_parts)))

            # Quality gate: question ≥ 10 chars, answer ≥ 20 chars
            if len(q_cleaned) >= 10 and len(a_cleaned) >= 20:
                q_lower = q_cleaned.lower()
                if q_lower not in seen_questions:
                    seen_questions.add(q_lower)
                    faqs.append({
                        "question": q_cleaned,
                        "answer": a_cleaned,
                        "source_url": source_url
                    })

    # =====================================================================
    # STEP 6: FALLBACK — regex-based extraction for unstructured PDFs
    # =====================================================================
    if len(faqs) < 2:
        faqs.clear()
        seen_questions.clear()
        qa_pattern = r'[Qq]\.?\s*([^\n?]+\?)\s*[Aa]\.?\s*([^\n]+(?:\n(?!Q\.)[^\n]+)*)'
        matches = re.findall(qa_pattern, text)
        for q, a in matches:
            # Filter line-by-line BEFORE clean_text replaces newlines
            filtered_parts = []
            for line in a.split('\n'):
                line_str = line.strip()
                if re.search(r'Page \d+ of \d+', line_str, re.I):
                    continue
                if re.search(r'^\d+ of \d+$', line_str):
                    continue
                if re.match(r'^\s*SEBI\s+FAQs?\s*$', line_str, re.I) or re.match(r'^\s*FAQs?\s+on\s+.*$', line_str, re.I):
                    continue
                filtered_parts.append(line_str)

            q_cleaned = clean_text(clean_question_text(q))
            a_cleaned = clean_text(clean_answer_text(" ".join(filtered_parts)))

            if len(q_cleaned) >= 10 and len(a_cleaned) >= 20:
                q_lower = q_cleaned.lower()
                if q_lower not in seen_questions:
                    seen_questions.add(q_lower)
                    faqs.append({
                        "question": q_cleaned,
                        "answer": a_cleaned,
                        "source_url": source_url
                    })

    return faqs


def scrape_faqs_from_pdf(pdf_url: str) -> List[Dict[str, Any]]:
    """
    Downloads a PDF and extracts questions and answers (FAQs).
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    try:
        response = requests.get(pdf_url, headers=headers, timeout=20)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"Failed to download PDF from {pdf_url}: {e}")
        raise RuntimeError(f"Download failed: {str(e)}")

    return scrape_faqs_from_pdf_bytes(response.content, pdf_url)


def find_column_index(headers: List[str], possible_names: List[str]) -> Optional[int]:
    """Finds index of column matching one of possible names (exact first, then substring)."""
    # 1. Exact case-insensitive match
    for name in possible_names:
        for idx, h in enumerate(headers):
            if h and str(h).strip().lower() == name.lower():
                return idx
    # 2. Case-insensitive substring match
    for name in possible_names:
        for idx, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return idx
    return None

import datetime
def parse_date(val: Any) -> Optional[datetime.datetime]:
    """Parses standard date strings and datetime objects from Excel/CSV."""
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime.combine(val, datetime.time.min)
    if not val or not isinstance(val, str):
        return None
    val = val.strip()
    # Try common formats in SEBI files
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None

def parse_excel_or_csv(file_contents: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parses spreadsheet row by row, matching headers and extracting values.
    Returns:
        List of dicts: {pdf_url, publication_date, topic, sub_topic, category}
    """
    rows = []
    headers = []
    data_rows = []

    if filename.lower().endswith('.csv'):
        try:
            text = file_contents.decode('utf-8')
        except UnicodeDecodeError:
            text = file_contents.decode('latin-1')
        reader = csv.reader(StringIO(text))
        csv_rows = list(reader)
        if not csv_rows:
            return []
        headers = [h.strip() for h in csv_rows[0]]
        data_rows = csv_rows[1:]
    else:
        wb = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)
        sheet = wb.active
        excel_rows = list(sheet.iter_rows(values_only=True))
        if not excel_rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in excel_rows[0]]
        data_rows = excel_rows[1:]

    # Map column headers to indices
    pdf_idx = find_column_index(headers, ['pdf links', 'pdf link', 'pdf_links', 'pdf_link', 'url', 'link'])
    date_idx = find_column_index(headers, ['dates', 'date', 'published date', 'publication date'])
    topic_idx = find_column_index(headers, ['topic', 'topics'])
    sub_topic_idx = find_column_index(headers, ['sub topic', 'sub-topic', 'subtopics', 'subtopic'])
    category_idx = find_column_index(headers, ['category', 'categories'])

    # We must have at least the PDF link column to process
    if pdf_idx is None:
        logger.error(f"Spreadsheet columns {headers} missing PDF URL column.")
        return []

    for r in data_rows:
        if not r or len(r) <= pdf_idx:
            continue
        pdf_url = r[pdf_idx]
        if not pdf_url:
            continue
        pdf_url = str(pdf_url).strip()
        # Simple URL format validate
        if not pdf_url.startswith(('http://', 'https://')):
            continue

        raw_date = r[date_idx] if (date_idx is not None and date_idx < len(r)) else None
        pub_date = parse_date(raw_date)

        topic = str(r[topic_idx]).strip() if (topic_idx is not None and topic_idx < len(r) and r[topic_idx] is not None) else None
        sub_topic = str(r[sub_topic_idx]).strip() if (sub_topic_idx is not None and sub_topic_idx < len(r) and r[sub_topic_idx] is not None) else None
        category = str(r[category_idx]).strip() if (category_idx is not None and category_idx < len(r) and r[category_idx] is not None) else None

        rows.append({
            "pdf_url": pdf_url,
            "document_publish_date": pub_date,
            "topic": topic,
            "sub_topic": sub_topic,
            "category": category
        })

    logger.info(f"Parsed {len(rows)} rows with PDF links from {filename}")
    return rows
